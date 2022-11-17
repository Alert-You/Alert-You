# 데이터 학교 데이터 정제
import pandas as pd
import openpyxl
from os import listdir
from asyncio.windows_events import NULL
from contextlib import nullcontext
import requests
import json

KEY = "780dacc6ed3443a2b47f4a3c27a6319c" # API key
AY = 2022
file = "학교기본정보_2022년10월31일기준.xlsx"
cur_wb = openpyxl.load_workbook(file, data_only=True)
print("===============reading===============")
cur_ws = cur_wb.active
tuple_set = set()
r = 2
region_dict = dict() # 학교 지역 정보를 담을 딕셔너리
address_dict = dict() # 학교 주소 정보를 담을 딕셔너리

while r <= 12534:
    apt_ofcdc_sc_code = cur_ws["A" + str(r)].value
    sd_schul_code = cur_ws["C" + str(r)].value
    tuple_set.add((apt_ofcdc_sc_code, sd_schul_code))
    region_dict[cur_ws["D" + str(r)].value] = cur_ws["G" + str(r)].value
    address_dict[cur_ws["D" + str(r)].value] = cur_ws["K" + str(r)].value # "(" + str(cur_ws["J" + str(r)].value) + ")" +
    r += 1

titles = ["학교명", "지역", "주소", "학년", "반"]

new_work_book = openpyxl.Workbook()
new_work_book.create_sheet("학교 정보")
new_work_sheet = new_work_book["학교 정보"]
del new_work_book["Sheet"]

for i in range(len(titles)):
    new_work_sheet[chr(65 + i) + "1"] = titles[i]

row = 2
for t in tuple_set:
    apt_ofcdc_sc_code, sd_schul_code = t
    req_url = f"https://open.neis.go.kr/hub/classInfo?KEY={KEY}&Type=json&pIndex=1&pSize=100&ATPT_OFCDC_SC_CODE={apt_ofcdc_sc_code}&SD_SCHUL_CODE={sd_schul_code}&AY={AY}"
    data = requests.get(req_url).json()
    classInfo = data.get("classInfo")
    if classInfo: # 에러가 없다면
        rowInfo = classInfo[1]["row"]
        print(rowInfo)
        for info in rowInfo:
            school_name = info.get('SCHUL_NM')
            region = region_dict.get(school_name)
            address = address_dict.get(school_name)
            grade = info.get('GRADE')
            class_num = info.get('CLASS_NM')
            new_work_sheet["A" + str(row)] = school_name
            new_work_sheet["B" + str(row)] = region
            new_work_sheet["C" + str(row)] = address
            new_work_sheet["D" + str(row)] = grade
            new_work_sheet["E" + str(row)] = class_num
            print(row)
            row += 1

new_work_book.save("데이터 정제 결과.xlsx")

